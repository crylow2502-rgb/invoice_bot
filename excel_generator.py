"""
Excel act generator for material transfer.
"""

import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


def generate_act(data: dict, template_path: str, output_path: str) -> str:
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path, keep_vba=False)
    ws = wb.active

    doc_number = data.get("doc_number", "")
    doc_date   = data.get("doc_date", "")
    supplier   = data.get("supplier", "")
    recipient  = data.get("recipient", "")
    items      = data.get("items", [])

    try:
        ws["A2"] = _parse_date(doc_date)
        ws["A2"].number_format = "DD.MM.YYYY"
    except Exception:
        pass

    try:
        ws["C11"] = supplier
        ws["C12"] = doc_number
        ws["C13"] = _parse_date(doc_date)
        ws["G7"]  = recipient
    except Exception:
        pass

    START_ROW = 16
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for i, item in enumerate(items):
        row = START_ROW + i

        name     = item.get("name", "")
        article  = item.get("article", "")
        quantity = item.get("quantity", 0)
        unit     = item.get("unit", "")

        full_name = name
        if article:
            full_name = f"{name}\nArt.: {article}"

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None

        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2).value = full_name
        ws.cell(row=row, column=2).alignment = left
        ws.cell(row=row, column=2).border = border
        ws.row_dimensions[row].height = 30 if article else 18

        ws.cell(row=row, column=5).value = quantity
        ws.cell(row=row, column=5).alignment = center
        ws.cell(row=row, column=5).border = border

        ws.cell(row=row, column=6).value = unit
        ws.cell(row=row, column=6).alignment = center
        ws.cell(row=row, column=6).border = border

    wb.save(output_path)
    return output_path


def _parse_date(date_str: str):
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except (ValueError, AttributeError):
            continue
    return date_str
